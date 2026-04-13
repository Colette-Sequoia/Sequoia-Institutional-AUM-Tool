#!/usr/bin/env python3
"""
Sequoia Investment Solutions - Institutional AUM Consolidation Tool
Backend server for processing 27Four institutional data files
"""

import os
import io
import json
import uuid
import traceback
import webbrowser
from datetime import date, datetime
from threading import Timer

from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
CORS(app)

# Global storage for session data
sessions = {}

# ============================================================================
# Utility Functions
# ============================================================================

def clean_string(s):
    """Clean and normalize string values"""
    if pd.isna(s) or s is None:
        return ""
    return str(s).strip()

def safe_float(val):
    """Convert to float, return 0 if invalid"""
    try:
        if pd.isna(val):
            return 0.0
        return float(val)
    except:
        return 0.0

def parse_date(date_val):
    """Parse various date formats to YYYY-MM-DD string"""
    if pd.isna(date_val):
        return None
    
    if isinstance(date_val, (datetime, pd.Timestamp)):
        return date_val.strftime('%Y-%m-%d')
    
    if isinstance(date_val, date):
        return date_val.strftime('%Y-%m-%d')
    
    # Try parsing string dates
    try:
        dt = pd.to_datetime(date_val)
        return dt.strftime('%Y-%m-%d')
    except:
        return str(date_val)

# ============================================================================
# Template Loading Functions
# ============================================================================

def load_template(file_bytes):
    """
    Load the institutional template file containing:
    - FUND MAP (LISPS NAMING → Fund Name)
    - PE REFERENCE (REFERENCE → PE + Retirement Fund Type)
    """
    try:
        df_fund_map = pd.read_excel(io.BytesIO(file_bytes), sheet_name='FUND MAP')
        df_pe_ref = pd.read_excel(io.BytesIO(file_bytes), sheet_name='PE REFERENCE')
        
        # Build fund lookup dictionary
        fund_map = {}
        for _, row in df_fund_map.iterrows():
            raw_name = clean_string(row.get('LISPS NAMING', '')).upper()
            if raw_name:
                fund_map[raw_name] = {
                    'fund_name': clean_string(row.get('Fund Name', '')),
                    'product': clean_string(row.get('Product', ''))
                }
        
        # Build PE reference lookup dictionary
        pe_map = {}
        for _, row in df_pe_ref.iterrows():
            ref_code = clean_string(row.get('REFERENCE', ''))
            if ref_code:
                pe_map[ref_code] = {
                    'pe': clean_string(row.get('PE', '')),
                    'retirement_fund_type': clean_string(row.get('Retirement Fund Type', ''))
                }
        
        return {
            'fund_map': fund_map,
            'pe_map': pe_map,
            'success': True
        }
    
    except Exception as e:
        return {
            'success': False,
            'error': f'Template load error: {str(e)}\n{traceback.format_exc()}'
        }

def load_mapping_file(file_bytes):
    """
    Load the institutional data mapping file containing:
    - EntityID Mapping (Raw Data → EntityID reference code)
    - Description Mapping (Raw Data → Standardized Description)
    """
    try:
        df_entity_map = pd.read_excel(io.BytesIO(file_bytes), sheet_name='EntityID Mapping')
        df_desc_map = pd.read_excel(io.BytesIO(file_bytes), sheet_name='Description Mapping')
        
        # Build entity ID lookup
        entity_map = {}
        for _, row in df_entity_map.iterrows():
            raw_id = row.get('Raw Data')
            entity_id = clean_string(row.get('EntityID', ''))
            if pd.notna(raw_id) and entity_id:
                # Store both as int and float keys to handle both formats
                entity_map[int(raw_id)] = entity_id
                entity_map[float(raw_id)] = entity_id
        
        # Build description lookup
        desc_map = {}
        for _, row in df_desc_map.iterrows():
            raw_desc = clean_string(row.get('Raw Data', '')).upper()
            std_desc = clean_string(row.get('Description', ''))
            if raw_desc:
                desc_map[raw_desc] = std_desc
        
        return {
            'entity_map': entity_map,
            'desc_map': desc_map,
            'success': True
        }
    
    except Exception as e:
        return {
            'success': False,
            'error': f'Mapping file load error: {str(e)}\n{traceback.format_exc()}'
        }

# ============================================================================
# Data Processing Functions
# ============================================================================

def process_27four_data(file_bytes, template_data, mapping_data, report_date, user_mappings=None):
    """
    Process the 27Four institutional data file
    
    Mapping flow:
    1. Entity ID → Reference Code (via mapping_data['entity_map'])
    2. Reference Code → PE + Retirement Fund Type (via template_data['pe_map'])
    3. Raw Fund → Standardized Description (via mapping_data['desc_map'])
    4. Standardized Description → Fund Name (via template_data['fund_map'])
    """
    if user_mappings is None:
        user_mappings = {'entity': {}, 'fund': {}}
    
    try:
        # Read the data file
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name='Sheet1')
        
        # Required columns
        required_cols = ['Date', 'Entity ID', 'Fund', 'Value', 'Name']
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            return {
                'success': False,
                'error': f'Missing required columns: {", ".join(missing)}'
            }
        
        results = []
        unmapped_entities = {}
        unmapped_funds = {}
        
        entity_map = mapping_data['entity_map']
        desc_map = mapping_data['desc_map']
        pe_map = template_data['pe_map']
        fund_map = template_data['fund_map']
        
        for idx, row in df.iterrows():
            try:
                # Extract raw values
                raw_entity_id = row.get('Entity ID')
                raw_fund = clean_string(row.get('Fund', '')).upper()
                aum_value = safe_float(row.get('Value', 0))
                entity_name = clean_string(row.get('Name', ''))
                
                # Skip rows with no AUM
                if aum_value == 0:
                    continue
                
                # === Entity ID Mapping ===
                ref_code = None
                pe = ""
                retirement_fund_type = ""
                
                # Check user mappings first
                if raw_entity_id in user_mappings['entity']:
                    mapping = user_mappings['entity'][raw_entity_id]
                    pe = mapping['pe']
                    retirement_fund_type = mapping['retirement_fund_type']
                else:
                    # Map Entity ID → Reference Code
                    ref_code = entity_map.get(raw_entity_id) or entity_map.get(int(raw_entity_id) if pd.notna(raw_entity_id) else None)
                    
                    if ref_code:
                        # Map Reference Code → PE + Retirement Fund Type
                        pe_info = pe_map.get(ref_code, {})
                        pe = pe_info.get('pe', '')
                        retirement_fund_type = pe_info.get('retirement_fund_type', '')
                    
                    # Track unmapped entities
                    if not pe or not retirement_fund_type:
                        if raw_entity_id not in unmapped_entities:
                            unmapped_entities[raw_entity_id] = {
                                'entity_id': raw_entity_id,
                                'entity_name': entity_name,
                                'ref_code': ref_code or 'NOT MAPPED',
                                'count': 0
                            }
                        unmapped_entities[raw_entity_id]['count'] += 1
                
                # === Fund Mapping ===
                fund_name = ""
                
                # Check user mappings first
                if raw_fund in user_mappings['fund']:
                    fund_name = user_mappings['fund'][raw_fund]['fund_name']
                else:
                    # Map Raw Fund → Standardized Description
                    std_desc = desc_map.get(raw_fund, '')
                    
                    if std_desc:
                        # Map Standardized Description → Fund Name
                        fund_info = fund_map.get(std_desc.upper(), {})
                        fund_name = fund_info.get('fund_name', '')
                    
                    # Track unmapped funds
                    if not fund_name:
                        if raw_fund not in unmapped_funds:
                            unmapped_funds[raw_fund] = {
                                'raw_fund': raw_fund,
                                'count': 0
                            }
                        unmapped_funds[raw_fund]['count'] += 1
                
                # Build result row
                result_row = {
                    'Date': report_date,
                    'Broker House Name': '',  # Always empty for institutional
                    'Broker Name': '',  # Always empty for institutional
                    'Retirement Fund Type': retirement_fund_type,
                    'Participating Employer': pe,
                    'Product': 'Institutional',  # Always "Institutional"
                    'LISP': 'NMG RFA',  # Always "NMG RFA"
                    'Fund Name': fund_name,
                    'InFlows (R)': 0,
                    'OutFlows (R)': 0,
                    'NetFlows (R)': 0,
                    'AUM (R)': aum_value
                }
                
                results.append(result_row)
            
            except Exception as e:
                print(f"Error processing row {idx}: {e}")
                continue
        
        # Create DataFrame
        df_result = pd.DataFrame(results)
        
        return {
            'success': True,
            'data': df_result,
            'unmapped_entities': list(unmapped_entities.values()),
            'unmapped_funds': list(unmapped_funds.values()),
            'row_count': len(df_result),
            'total_aum': df_result['AUM (R)'].sum() if len(df_result) > 0 else 0
        }
    
    except Exception as e:
        return {
            'success': False,
            'error': f'Processing error: {str(e)}\n{traceback.format_exc()}'
        }

# ============================================================================
# Excel Output Functions
# ============================================================================

def create_excel_output(df_data, report_date):
    """Create formatted Excel output matching retail format"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Main Worksheet"
    
    # Define styles
    header_font = Font(name='Calibri Light', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Calibri Light', size=12)
    data_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    
    border_side = Side(style='thin', color='000000')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Row shading colors
    light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Write headers
    headers = list(df_data.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows with alternating shading
    for row_idx, row_data in enumerate(df_data.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = border
            
            # Apply number formatting for currency columns
            col_name = headers[col_idx - 1]
            if '(R)' in col_name:
                cell.number_format = '#,##0.00'
                cell.alignment = number_alignment
            else:
                cell.alignment = data_alignment
            
            # Alternating row shading
            if row_idx % 2 == 0:
                cell.fill = light_fill
    
    # Add totals row
    totals_row = len(df_data) + 2
    ws.cell(row=totals_row, column=1, value='TOTAL')
    ws.cell(row=totals_row, column=1).font = Font(name='Calibri Light', size=12, bold=True)
    
    # Calculate totals for currency columns (plain values, not formulas)
    for col_idx, col_name in enumerate(headers, 1):
        if '(R)' in col_name:
            total_value = df_data[col_name].sum()
            cell = ws.cell(row=totals_row, column=col_idx, value=total_value)
            cell.font = Font(name='Calibri Light', size=12, bold=True)
            cell.number_format = '#,##0.00'
            cell.alignment = number_alignment
            cell.border = border
    
    # Set column widths
    column_widths = {
        'Date': 12,
        'Broker House Name': 25,
        'Broker Name': 25,
        'Retirement Fund Type': 30,
        'Participating Employer': 35,
        'Product': 15,
        'LISP': 12,
        'Fund Name': 40,
        'InFlows (R)': 15,
        'OutFlows (R)': 15,
        'NetFlows (R)': 15,
        'AUM (R)': 18
    }
    
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = column_widths.get(header, 15)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Enable auto-filter
    ws.auto_filter.ref = f'A1:{chr(64 + len(headers))}{len(df_data) + 1}'
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# ============================================================================
# Flask Routes
# ============================================================================

@app.route('/')
def index():
    """Serve the HTML frontend"""
    return send_from_directory('.', 'aum_insto.html')

@app.route('/api/new-session', methods=['POST'])
def new_session():
    """Create a new processing session"""
    session_id = str(uuid.uuid4())
    sessions[session_id] = {
        'template': None,
        'mapping': None,
        'data_file': None,
        'processed_data': None,
        'user_mappings': {'entity': {}, 'fund': {}}
    }
    return jsonify({'session_id': session_id})

@app.route('/api/upload-template', methods=['POST'])
def upload_template():
    """Upload and parse the institutional template file"""
    try:
        session_id = request.form.get('session_id')
        if not session_id or session_id not in sessions:
            return jsonify({'success': False, 'error': 'Invalid session'})
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'})
        
        file = request.files['file']
        file_bytes = file.read()
        
        result = load_template(file_bytes)
        
        if result['success']:
            sessions[session_id]['template'] = result
            return jsonify({
                'success': True,
                'fund_count': len(result['fund_map']),
                'pe_count': len(result['pe_map'])
            })
        else:
            return jsonify(result)
    
    except Exception as e:
        return jsonify({'success': False, 'error': f'Upload error: {str(e)}'})

@app.route('/api/upload-mapping', methods=['POST'])
def upload_mapping():
    """Upload and parse the institutional data mapping file"""
    try:
        session_id = request.form.get('session_id')
        if not session_id or session_id not in sessions:
            return jsonify({'success': False, 'error': 'Invalid session'})
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'})
        
        file = request.files['file']
        file_bytes = file.read()
        
        result = load_mapping_file(file_bytes)
        
        if result['success']:
            sessions[session_id]['mapping'] = result
            return jsonify({
                'success': True,
                'entity_count': len(result['entity_map']),
                'desc_count': len(result['desc_map'])
            })
        else:
            return jsonify(result)
    
    except Exception as e:
        return jsonify({'success': False, 'error': f'Upload error: {str(e)}'})

@app.route('/api/upload-data', methods=['POST'])
def upload_data():
    """Upload the 27Four data file"""
    try:
        session_id = request.form.get('session_id')
        if not session_id or session_id not in sessions:
            return jsonify({'success': False, 'error': 'Invalid session'})
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'})
        
        file = request.files['file']
        sessions[session_id]['data_file'] = file.read()
        
        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'error': f'Upload error: {str(e)}'})

@app.route('/api/process', methods=['POST'])
def process_data():
    """Process the uploaded data file"""
    try:
        data = request.json
        session_id = data.get('session_id')
        report_date = data.get('report_date')
        
        if not session_id or session_id not in sessions:
            return jsonify({'success': False, 'error': 'Invalid session'})
        
        session = sessions[session_id]
        
        if not session['template']:
            return jsonify({'success': False, 'error': 'Template not uploaded'})
        
        if not session['mapping']:
            return jsonify({'success': False, 'error': 'Mapping file not uploaded'})
        
        if not session['data_file']:
            return jsonify({'success': False, 'error': 'Data file not uploaded'})
        
        # Process the data
        result = process_27four_data(
            session['data_file'],
            session['template'],
            session['mapping'],
            report_date,
            session['user_mappings']
        )
        
        if result['success']:
            session['processed_data'] = result['data']
            
            return jsonify({
                'success': True,
                'row_count': result['row_count'],
                'total_aum': result['total_aum'],
                'unmapped_entities': result['unmapped_entities'],
                'unmapped_funds': result['unmapped_funds']
            })
        else:
            return jsonify(result)
    
    except Exception as e:
        return jsonify({'success': False, 'error': f'Processing error: {str(e)}\n{traceback.format_exc()}'})

@app.route('/api/save-mapping', methods=['POST'])
def save_mapping():
    """Save user-provided mapping for unmapped items"""
    try:
        data = request.json
        session_id = data.get('session_id')
        mapping_type = data.get('type')  # 'entity' or 'fund'
        key = data.get('key')
        mapping_data = data.get('mapping')
        
        if not session_id or session_id not in sessions:
            return jsonify({'success': False, 'error': 'Invalid session'})
        
        session = sessions[session_id]
        
        if mapping_type == 'entity':
            session['user_mappings']['entity'][key] = mapping_data
        elif mapping_type == 'fund':
            session['user_mappings']['fund'][key] = mapping_data
        else:
            return jsonify({'success': False, 'error': 'Invalid mapping type'})
        
        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'error': f'Save error: {str(e)}'})

@app.route('/api/download', methods=['POST'])
def download_excel():
    """Generate and download the Excel output"""
    try:
        data = request.json
        session_id = data.get('session_id')
        report_date = data.get('report_date')
        
        if not session_id or session_id not in sessions:
            return jsonify({'success': False, 'error': 'Invalid session'})
        
        session = sessions[session_id]
        
        if session['processed_data'] is None or len(session['processed_data']) == 0:
            return jsonify({'success': False, 'error': 'No processed data available'})
        
        # Create Excel file
        excel_output = create_excel_output(session['processed_data'], report_date)
        
        # Generate filename
        date_str = report_date.replace('-', '')[:6]  # YYYYMM format
        filename = f'AUM_Institutional_{date_str}.xlsx'
        
        return send_file(
            excel_output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'success': False, 'error': f'Download error: {str(e)}\n{traceback.format_exc()}'})

# ============================================================================
# Main
# ============================================================================

if __name__ == '__main__':
    # Get port from environment variable (Render) or use 5051 for local
    port = int(os.environ.get('PORT', 5051))
    
    # Only open browser when running locally
    if os.environ.get('PORT') is None:
        def open_browser():
            webbrowser.open(f'http://localhost:{port}')
        Timer(2, open_browser).start()
    
    print("\n" + "=" * 50)
    print("  Sequoia Capital Management")
    print("  Institutional AUM Consolidation Tool")
    print("=" * 50)
    print(f"\n  Server running on port {port}")
    print("  Keep this window open while using the tool.")
    print("  Press Ctrl+C to stop.")
    print("\n" + "=" * 50 + "\n")
    
    app.run(host='0.0.0.0', port=port, debug=False)
